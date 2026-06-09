"""
panelDistri — Flask Application
Panel de visualización de pedidos para TV + descarga de reporte diario.

Rutas:
  /             → redirige a /tv-panel
  /tv-panel     → panel en vivo para televisor
  /reporte      → página de descarga de informe
  /api/pedidos/estado      → JSON con pedidos, stats y motivos (polling cada 10s)
  /api/pedido/meta         → POST actualizar metadata de un pedido
  /api/motivo              → POST registrar motivo de orden incompleta
  /api/motivo/<id>/borrar  → POST eliminar un motivo
  /api/reporte/descargar   → GET descargar Excel del día
"""

from flask import Flask, render_template, jsonify, send_file, request, redirect, url_for
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

import db

app = Flask(__name__)
app.secret_key = 'paneldistri-2024-secret'


# ─── Inicialización ───────────────────────────────────────────────────────────

with app.app_context():
    db.init_panel_tables()


# ─── Rutas principales ────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('tv_panel'))


@app.route('/tv-panel')
def tv_panel():
    return render_template('tv_panel.html')


@app.route('/reporte')
def reporte():
    return render_template('reporte.html')


# ─── API JSON ─────────────────────────────────────────────────────────────────

@app.route('/api/pedidos/estado')
def api_estado():
    pedidos = db.get_pedidos_estado()
    stats   = db.get_stats_hoy()
    motivos = db.get_motivos_hoy()
    return jsonify({
        'pedidos':   pedidos,
        'stats':     stats,
        'motivos':   motivos,
        'timestamp': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
        'fecha':     date.today().strftime('%d/%m/%Y'),
    })


@app.route('/api/pedido/meta', methods=['POST'])
def api_pedido_meta():
    data = request.get_json(silent=True) or {}
    pedido = data.get('pedido', '').strip()
    if not pedido:
        return jsonify({'error': 'Campo "pedido" requerido'}), 400

    ok = db.upsert_pedido_meta(
        pedido       = pedido,
        prioridad    = data.get('prioridad'),
        pickea       = data.get('pickea'),
        corre        = data.get('corre'),
        confirmado   = data.get('confirmado'),
        fecha_pedido = data.get('fecha_pedido'),
        cod_cliente  = data.get('cod_cliente'),
    )
    if ok:
        return jsonify({'ok': True})
    return jsonify({'error': 'Error al guardar en la base de datos'}), 500


@app.route('/api/motivo', methods=['POST'])
def api_motivo():
    data = request.get_json(silent=True) or {}
    pedido = data.get('pedido', '').strip()
    motivo = data.get('motivo', '').strip()
    if not pedido or not motivo:
        return jsonify({'error': 'Campos "pedido" y "motivo" requeridos'}), 400

    nuevo_id = db.insertar_motivo(
        pedido        = pedido,
        motivo        = motivo,
        cliente       = data.get('cliente'),
        submotivo     = data.get('submotivo'),
        observaciones = data.get('observaciones'),
        registrado_por= data.get('registrado_por'),
    )
    if nuevo_id:
        return jsonify({'ok': True, 'id': nuevo_id})
    return jsonify({'error': 'Error al guardar motivo'}), 500


@app.route('/api/motivo/<int:motivo_id>/borrar', methods=['POST'])
def api_motivo_borrar(motivo_id):
    ok = db.eliminar_motivo(motivo_id)
    if ok:
        return jsonify({'ok': True})
    return jsonify({'error': 'Error al eliminar motivo'}), 500


# ─── Descarga de reporte ──────────────────────────────────────────────────────

@app.route('/api/reporte/descargar')
def api_reporte_descargar():
    fecha_str = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))

    try:
        pedidos = db.get_pedidos_estado()
        stats   = db.get_stats_hoy()
        motivos = db.get_motivos_hoy()

        wb = openpyxl.Workbook()

        header_fill  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font  = Font(color='FFFFFF', bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        alta_fill    = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        media_fill   = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
        baja_fill    = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        ok_fill      = PatternFill(start_color='375623', end_color='375623', fill_type='solid')

        # ── Hoja 1: Pedidos ──────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Pedidos'
        ws1.row_dimensions[1].height = 22

        headers1 = [
            'Pedido', 'Cliente', 'Cód.Cliente', 'Fecha', 'Prioridad',
            'Pickea', 'Corre', 'SS.Pick', 'SS.Corre', 'Estado', 'Confirmado'
        ]
        ws1.append(headers1)
        for cell in ws1[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align

        for p in pedidos:
            ss_pick  = f"{p.get('ss_pick', 0)}/{p.get('ss_pick_total', 0)}"
            ss_corre = f"{p.get('ss_corre', 0)}/{p.get('ss_corre_total', 0)}"
            row = [
                p['pedido'], p['cliente'], p['cod_cliente'], p['fecha_pedido'],
                p['prioridad'], p['pickea'], p['corre'],
                ss_pick, ss_corre, p['estado'],
                'Sí' if p['confirmado'] else 'No',
            ]
            ws1.append(row)
            # Color de prioridad
            prio = p['prioridad'].upper()
            fill = alta_fill if prio == 'ALTA' else media_fill if prio == 'MEDIA' else baja_fill
            ws1.cell(row=ws1.max_row, column=5).fill = fill
            ws1.cell(row=ws1.max_row, column=5).font = Font(color='FFFFFF', bold=True)
            # Color confirmado
            if p['confirmado']:
                ws1.cell(row=ws1.max_row, column=11).fill = ok_fill
                ws1.cell(row=ws1.max_row, column=11).font = Font(color='FFFFFF')

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 12), 40)

        # ── Hoja 2: Informe ──────────────────────────────────────────────────
        ws2 = wb.create_sheet('Informe')
        ws2.append(['INFORME DEL DÍA', fecha_str])
        ws2['A1'].font = Font(bold=True, size=14)
        ws2.append([])
        ws2.append(['Métrica', 'Valor'])
        ws2['A3'].font = header_font
        ws2['B3'].font = header_font
        ws2['A3'].fill = header_fill
        ws2['B3'].fill = header_fill
        for metric, val in [
            ('Órdenes Activas (con control hoy)',  stats['ordenes_activas']),
            ('Finalizadas hoy',                    stats['finalizadas_hoy']),
            ('En control ahora mismo',             stats['en_control']),
            ('Total en cola',                      stats['total_cola']),
        ]:
            ws2.append([metric, val])
        ws2.column_dimensions['A'].width = 38
        ws2.column_dimensions['B'].width = 14

        # ── Hoja 3: Motivos ──────────────────────────────────────────────────
        ws3 = wb.create_sheet('Motivos Incompletas')
        headers3 = ['Pedido', 'Cliente', 'Motivo', 'Submotivo', 'Observaciones', 'Registrado Por', 'Hora']
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
        for m in motivos:
            ws3.append([
                m['pedido'], m['cliente'], m['motivo'], m['submotivo'],
                m['observaciones'], m['registrado_por'], m['hora'],
            ])
        for col in ws3.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws3.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 12), 45)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"reporte_distri_{fecha_str}.xlsx"
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ─── Inicio ───────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5003)
